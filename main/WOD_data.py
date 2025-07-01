# Imports packages
import openpyxl 
from datetime import datetime, time
import os
import csv

# Defines class ETL
class ETL:
    def __init__(self):
        print("Beginning ETL Process")
        self.file_selector()
        self.file_loader()
        self.setup_dict()
        self.ext_trans()
        self.ext_trans_time()
        self.setup_datetime_ind()
        self.load_to_csv_D1()
        self.load_to_csv_D05()
        print("Completed ETL Process")
        return
    def file_selector(self):
        # Opens file directory to search for input WOD xlsx file
        print(f"Files in Directory WOD")
        for file_item in os.listdir("../../Data/TELEOS_1/WOD"):
            print(file_item)
        self.input_file_name = input(f"Input WOD Data file name with xlsx extension: ")
        self.input_file_path = f"../../Data/TELEOS_1/WOD/{self.input_file_name}"
        print("Input file path:", self.input_file_path)
        return
    def file_loader(self):
        # Loads workbook
        self.wb = openpyxl.load_workbook(self.input_file_path)
    def setup_dict(self):
        # Sets up dictionaries and lists
        self.dict_sheets = {"pos_x":"0x80070000 - Data Order 1",\
                            "pos_y":"0x80080000 - Data Order 1",\
                            "pos_z":"0x80090000 - Data Order 1",\
                            "vel_x":"0x800A0000 - Data Order 1",\
                            "vel_y":"0x800B0000 - Data Order 1",\
                            "vel_z":"0x800C0000 - Data Order 1"}
        self.dict_lists = {"datetime":[],\
                            "pos_x":[],\
                            "pos_y":[],\
                            "pos_z":[],\
                            "vel_x":[],\
                            "vel_y":[],\
                            "vel_z":[]}
        self.dict_trans = {"pos_x":self.pos_transform,\
                            "pos_y":self.pos_transform,\
                            "pos_z":self.pos_transform,\
                            "vel_x":self.vel_transform,\
                            "vel_y":self.vel_transform,\
                            "vel_z":self.vel_transform }
        self.dict_dt_ind = {}
    # Transforms position data units of dm to m
    def pos_transform(self, data_raw):
        data_trans = data_raw/10
        return data_trans
    # Transforms velocity data units of mm/s to m/s
    def vel_transform(self, data_raw):
        data_trans = data_raw/1000
        return data_trans
    # Loops through sheets, extracts and transforms raw data to SI unit data
    def ext_trans(self):
        for field in self.dict_sheets:
            trans_mtd = self.dict_trans[field]
            self.ws = self.wb[self.dict_sheets[field]]
            print(f"For Data {field}")
            print('Total number of rows: '+str(self.ws.max_row)+'. And total number of columns: '+str(self.ws.max_column)+'\n')
            for row_num in range(11, 2821+1):
                val_raw = self.ws.cell(row=row_num, column=1).value
                val_trans = trans_mtd(val_raw)
                self.dict_lists[field].append(val_trans)
        print("Extracted and transformed all PV data\n")
        return
    # Extracts and transforms datetime from input file 
    def ext_trans_time(self):
        self.ws = self.wb[self.dict_sheets["pos_x"]]
        print("For Data datetime")
        print('Total number of rows: '+str(self.ws.max_row)+'. And total number of columns: '+str(self.ws.max_column)+'\n')
        for row_num in range(11, 2821+1):
            val_raw = self.ws.cell(row=row_num, column=8).value
            self.dict_lists["datetime"].append(val_raw)
        self.num_datapoints = len(self.dict_lists["datetime"])
        print("Extracted all datetime data\n")
        return
    # Generates stop indices for datetime list seperating data based on day
    def setup_datetime_ind(self):
        dt_prev  = self.dict_lists["datetime"][0]
        # Loops through indices of datetime list
        for ind in range(self.num_datapoints):    
            dt = self.dict_lists["datetime"][ind]
            if dt.date() != dt_prev.date():
                self.dict_dt_ind[ind] = dt_prev.strftime("%Y%m%d")
                dt_prev = dt
        self.dict_dt_ind[self.num_datapoints] = self.dict_lists["datetime"][-1].strftime("%Y%m%d")
        print(f"Number of days of data: {len(self.dict_dt_ind)}\n")
        return
    # Loads data to output csv files for each day's worth of WOD GPS data
    def load_to_csv_D1(self):
        print("Loading daily data to csv files")
        self.output_file_directory = "../../Data/TELEOS_1/WOD_proc"
        # Sets up first start index
        ind_start = 0
        # Loops through stop indices for datetime list
        for ind_stop in self.dict_dt_ind:
            # Formats date into output file name
            date = self.dict_dt_ind[ind_stop]
            self.output_file_name = f"WOD_GPS_Test_Data_PVT_{date}.csv"
            self.output_file_path = f"{self.output_file_directory}/{self.output_file_name}"
            print(f"Writing data to csv file {self.output_file_name}")
            # Writes data to output file
            with open(self.output_file_path, "w", newline="") as file:
                writer = csv.writer(file)
                writer.writerow(["Time Stamp",\
                                "X (m)",\
                                "Y (m)",\
                                "Z (m)",\
                                "Vel X (m/s)",\
                                "Vel Y (m/s)",\
                                "Vel Z (m/s)"])
                for ind in range(ind_start, ind_stop):
                    row_entry = []
                    for field in self.dict_lists:
                        row_entry.append(self.dict_lists[field][ind])
                    writer.writerow(row_entry)
            ind_start = ind_stop
        print("Finished loading daily data to csv files")
        return
    # Loads data to output csv files for each day's half day worth of WOD GPS data
    def load_to_csv_D05(self):
        print("Loading half day of data to csv files")
        self.output_file_directory = "../../Data/TELEOS_1/WOD_proc"
        # Sets up first start index
        ind_start = 0
        time_cutoff = time(12)
        # Loops through stop indices for datetime list
        for ind_stop in self.dict_dt_ind:
            # Formats date into output file name
            date = self.dict_dt_ind[ind_stop]
            self.output_file_name = f"WOD_GPS_Test_Data_PVT_{date}_0.5.csv"
            self.output_file_path = f"{self.output_file_directory}/{self.output_file_name}"
            print(f"Writing data to csv file {self.output_file_name}")
            # Writes data to output file
            with open(self.output_file_path, "w", newline="") as file:
                writer = csv.writer(file)
                writer.writerow(["Time Stamp",\
                                "X (m)",\
                                "Y (m)",\
                                "Z (m)",\
                                "Vel X (m/s)",\
                                "Vel Y (m/s)",\
                                "Vel Z (m/s)"])
                # Conditional check if data starts from before UTC 12
                if self.dict_lists["datetime"][ind_start].time() < time_cutoff\
                    and self.dict_lists["datetime"][ind_stop-1].time() > time_cutoff:
                    for ind in range(ind_start, ind_stop):    
                        if self.dict_lists["datetime"][ind].time() > time_cutoff:
                            ind_half_stop = ind
                            break
                else:
                    ind_half_stop = ind_stop
                # Extracts only half day worth of data
                for ind in range(ind_start, ind_half_stop):
                    row_entry = []
                    for field in self.dict_lists:
                        row_entry.append(self.dict_lists[field][ind])
                    writer.writerow(row_entry)
            ind_start = ind_stop
        print("Finished loading half day of data to csv files")
        return

# Creates ETL object which executes ETL process
test = ETL()